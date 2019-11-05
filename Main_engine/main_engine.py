#coding:utf-8
import hashlib
import json
import django
import timeit
import os
from multiprocessing import Process, Queue, Manager
from collections import OrderedDict
import pefile
django.setup()

from Main_engine.Extract_Engine import pe2idb
from Main_engine.Extract_Engine.Flowchart_feature import extract_asm_and_const
from Main_engine.Extract_Engine.PE_feature import extract_pe
from Main_engine.Analzer_Engine import analyze_pe, analyze_flowchart
from openpyxl import load_workbook, Workbook
from Main_engine.Check_Packing import Packer_Detect2
from Main_engine.Unpacking import unpack_module
from Main_engine.models import *

idb_file_path = "C:\\malware\\all_result\\idb\\"
pe_file_path = "C:\\malware\\all_result\\pe\\"

class Pe_Files_Check:
    '''
        * get_unique_pe_list
        -. idb로 변환할 PE 파일 중복 확인
        -. 파일명:해시값 딕셔너리에 저장 -> db or file
        -. 중복된 파일 제거
        -. Check the completely same PE file.
        -. make a dictionary of filename(key) and hashvalue(value)
        -. And then remove the same PE file.
        # exe_q에 idb로 변환할 exe파일을 쌓는다
        exe_q=Queue()
        * Return value
           dictionary of filename(key) and hashvalue(value)
        '''
    def __init__(self,pe_dir_path):
        self.pe_dir_path = pe_dir_path
        self.idat = 0
        self.idat64 = 1
        self.pe_check_error = -1
        self.pe_hash_dict = {}
        self.HEX_M_32 = 0x14c
        self.HEX_M_64 = 0x200

    def get_unique_pe_list(self):
        exe_list = os.listdir(self.pe_dir_path)
        for f in exe_list:
            f_path = os.path.join(self.pe_dir_path, f)
            f_hash = hashlib.sha256(open(f_path, 'rb').read()).hexdigest()

            # file hash 중복 = 완전히 같은 파일
            # 해당 파일은 삭제(이미 diffing할 동일 파일이 존재하므로)
            if f_hash in self.pe_hash_dict.values():
                os.remove(f_path)
            else:
                #os.rename(f_path, os.path.join(self.pe_dir_path, f_hash))
                print('a')

                # 파일은 삭제하지만 해당 파일명(절대경로)와 해시정보는 DB에 있어야함.
            # 추후 시각화할 때 정보 필요
            self.pe_hash_dict[f_path] = f_hash


        # 이후에는 DB에 저장.
        # dictionary로 넘겨서 self.pe_hash_dict.value()의 유니크한 값들만 idb로 변환.
        # 나중에 서버에 올린 후에는 받은 파일 중 완전히 같은 파일은 서버 파일시스템에 저장하지 않는게 좋을 것 같다.
        # 예를 들어, file1, file2의 해시값이 동일한 경우
        # DB에는 file1과 file2의 파일명 그리고 해시값을 저장한다.
        # 그리고 서버 파일시스템 내에서 둘 중 하나의 파일은 지운다.
        # 일단 더 먼저 나오는 파일을 살리고 아닌 파일은 삭제하도록 하였다. 어차피 동일 파일이니깐.
        with open(r"C:\malware\all_result\test_pelist.txt", 'w') as pelist:
            json.dump(self.pe_hash_dict, pelist, ensure_ascii=False, indent='\t')

        return self.pe_hash_dict

    def unpack_pe(self):
        Packer_Detect2.sample_packer_type_detect(self.pe_dir_path)

        sample_folder_path = self.pe_dir_path
        save_folder_path = r"C:\malware\packing_info"
        pack_path = os.path.join(save_folder_path, 'packed')
        unpack_path = os.path.join(save_folder_path, 'unpacked')
        if not (os.path.isdir(save_folder_path)): os.makedirs(save_folder_path)
        if not (os.path.isdir(pack_path)): os.makedirs(pack_path)
        if not (os.path.isdir(unpack_path)): os.makedirs(unpack_path)

        queue = unpack_module.mains(sample_folder_path)

        proc_list = []
        for _ in range(0, 5):
            proc = Process(target=unpack_module.packer_check, args=(queue, pack_path, unpack_path,))
            proc_list.append(proc)
        for proc in proc_list:
            proc.start()
        for proc in proc_list:
            proc.join()

def convert_idb(PATH,IDB_PATH):
    # idb 변환
    return pe2idb.create_idb(PATH, IDB_PATH)

def multiprocess_file(q, return_dict, flag):

    while q.empty() != True:
        f_path = q.get()

        if flag == 'idb':
            # 여기에 조건문 하나 더 추가해서 바로 idb 추출하는게 아니라 db에서 이미 뽑힌 정보 있는지 확인하고
            # 저장된게 있으면 해당 파일 정보 dict의 경로를 db에서 가져와서
            # json.load로 읽어서 dict를 받음

            # if db 미 존재
            file_filter = f_path[f_path.rfind('\\')+1:-4]

            try:
                file = Filter.objects.get(filehash=file_filter)
                fd1 = open(file.idb_filepath + ".txt","rb").read()
                info = json.loads(fd1, encoding='utf-8')
                print('idb존재함')
            except:
                info = extract_asm_and_const.basicblock_idb_info_extraction(f_path)  # 함수대표값 및 상수값 출력
                with open(r"C:\malware\all_result\idb" + "\\" + file_filter + ".txt", 'w') as makefile:
                    json.dump(info, makefile, ensure_ascii=False, indent='\t')
                Filter.objects.create(filehash=info['file_name'], idb_filepath=idb_file_path+file_filter)
                print('idb없음')

        elif flag == 'pe':

            file_filter2 = f_path[f_path.rfind('\\') + 1:]

            try:
                print('zzzzzz1')
                print(file_filter2)
                pe_file = Filter.objects.get(filehash=file_filter2)
                print('ssdfasdfasdf')
                print(pe_file.pe_filepath)
                if pe_file.pe_filepath is not None:
                    fd1 = open(pe_file.pe_filepath + ".txt", "rb").read()
                    info = json.loads(fd1, encoding='utf-8')
                    print('pe존재함')
                else:
                    try:
                        pe = pefile.PE(f_path)
                        print('9999999999999999999999999999')
                        info = extract_pe.Pe_Feature(f_path, pe).all()  # pe 속성 출력
                        with open(r"C:\malware\all_result\pe" + "\\" + file_filter2 + ".txt", 'w') as makefile:
                            json.dump(info, makefile, ensure_ascii=False, indent='\t')
                        print('ads')
                        #pe_file.pe_filepath(pe_file_path + file_filter2)
                        #pefile.save()
                        print('pe없음')
                    except:
                        print('pe error !')
                        continue
            except:
                try:
                    pe = pefile.PE(f_path)
                    info = extract_pe.Pe_Feature(f_path, pe).all()  # pe 속성 출력
                    with open(r"C:\malware\all_result\pe" + "\\" + file_filter2 + ".txt", 'w') as makefile:
                        json.dump(info, makefile, ensure_ascii=False, indent='\t')
                    tmp = Filter.objects.get(filehash=file_filter2)
                    #tmp.update(pe_filepath=pe_file_path + file_filter2)
                    print('없음')
                except:
                    print('pe error !')
                    continue

        return_dict[f_path] = info


class Exract_Feature:
    def __init__(self, path, idb_path):
        self.path = path
        self.idb_path = idb_path

    def export_by_multi(self, flag):

        if flag == 'idb':
            path = self.idb_path
        elif flag == 'pe':
            path = self.path

        try:
            q = Queue()
            manager = Manager()
            pe2idb.exe_list_to_queue(path, q)
            return_dict = manager.dict()
            procs = list()
            for i in range(os.cpu_count() // 2 + 1):
                proc = Process(target=multiprocess_file, args=[q, return_dict, flag])
                procs.append(proc)
                proc.start()
            for p in procs:
                p.join()
            return return_dict
        except:
            return False

    def export_idb_info(self, flag):

        export_idb = self.export_by_multi(flag)

        if export_idb != False:
            return export_idb
        else:
            return False

    def export_pe_info(self, flag):

        export_pe = self.export_by_multi(flag)

        if export_pe != False:

            for dict_list in export_pe.values():

                # 파일 저장은 우리를 위한 임시적인 행위
                with open(r"C:\malware\all_result\pe"+"\\" + dict_list['file_name'] + ".txt", 'w') as makefile:
                    json.dump(dict_list, makefile, ensure_ascii=False, indent='\t')

                # 실제론 db 저장(?)
                #
                #
            return export_pe
        else:
            return False

class Analyze_files:
    def __init__(self, all_idb_info, all_pe_info):
        self.all_pe_info = all_pe_info
        self.all_idb_info = all_idb_info

    def calculate_heuristic(self, idb_result, pe_result):
        '''
                가중치가 부여된 점수들을 더해서 반환해주는 함수
                *다 더했을 때 최대나 최소안에 있는지 확인하는 로직을 넣어주고 예외처리 해주면 될 듯
                :return: final score
                '''
        # 최종 휴리스틱 스코어

        real_final = OrderedDict()

        for key_i, key_pe in zip(idb_result.items(), pe_result.items()):
            idb_final_score = OrderedDict()
            pe_final_score = OrderedDict()
            for value_i, value_pe in zip(key_i[1].items(), key_pe[1].items()):
                semifinal = [0, 0, 0, 0, 0, 0, 0, 0, 0]
                semifinal[0] = (value_pe[1]['file_hash'])
                semifinal[1] = (value_pe[1]['time_date_stamp'])
                semifinal[2] = (value_i[1]['bbh'])
                semifinal[3] = (value_i[1]['const_value'])
                semifinal[4] = (value_pe[1]['section_score'])
                semifinal[5] = (value_pe[1]['auth_score'])
                semifinal[6] = (value_pe[1]['pdb_score'])
                semifinal[7] = (value_pe[1]['imphash'])
                semifinal[8] = (value_pe[1]['rich'])

                idb_final_score[value_i[0]] = semifinal
                pe_final_score[value_pe[0]] = semifinal

            real_final[key_i[0]] = idb_final_score
            real_final[key_pe[0]] = pe_final_score

        return real_final

    def analyze_idb(self, yun_sorted_pe):
        idb = analyze_flowchart.AnalyzeFlowchart(self.all_idb_info)
        idb_result, yun_all = idb.analyze_all(yun_sorted_pe)

        return idb_result, yun_all

    def analyze_pe(self):
        pe = analyze_pe.AnalyzePE(self.all_pe_info)
        pe_split = pe.pe_parser()
        pe_result, yun_pe = pe.analyze_all(pe_split)

        return pe_result, yun_pe

'''
    total score to the excel file
'''
def out_xlsx(path, result_dict):
    try:
        wb = load_workbook(path)
    except:
        wb = Workbook()
    ws = wb.create_sheet()
    ws = wb.active

    ws.title = 'result_xlsx'
    title = ['BASE_FILE', 'COMP_FILE', 'FILE HASH', 'TIME STAMP', 'BB HASH', 'CONSTANT', 'SECTION', 'AUTH', 'PDB', 'IMPORT HASH', 'RICH']
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']

    for i in range(len(title)):
        ws[f'{cols[i]}1'] = title[i]
    target_count = len(result_dict) - 1
    start_row_num = 2

    for base, targets in result_dict.items():
        current_row_num = start_row_num
        ws[f'A{current_row_num}'] = base
        for t_name, t_infos in targets.items():
            ws[f'B{current_row_num}'] = t_name
            current_info = 0
            for t_info in t_infos:
                ws[f'{cols[current_info + 2]}{current_row_num}'] = t_info
                current_info += 1
            current_row_num += 1
        ws.merge_cells(f"A{start_row_num}:A{current_row_num - 1}")
        current_row_num += 1
        start_row_num = current_row_num

    #    wb.remove(wb['Sheet1'])
    wb.save(path)



def start_engine():
    '''
    웹 서버에서 메인 엔진을 호출하면 엔진을 돌리기위한 함수
    * 백앤드 엔진에서는 사용되지 않음
    * 지금은 서버 테스트만을 위해서 만든 것이므로 무시
    '''
    PATH = r"C:\malware\mal_exe"
    IDB_PATH = r"C:\malware\mal_idb"
    RESUT_IDB_PATH = r"C:\malware\all_result_idb"
    RESUT_PE_PATH = r"C:\malware\all_result_pe"

    # 1. pe 해시 체크 (동일한 파일 필터), 2.패킹 체크
    pe_check = Pe_Files_Check(PATH)
    file_hash_dict = pe_check.get_unique_pe_list()

    # 3. pe파일(+패킹 체크) -> idb 변환
    flag = convert_idb(PATH, IDB_PATH)
    Features = Exract_Feature(PATH, IDB_PATH)

    # 4. 정보 추출(idb,pe)
    if flag == True:
        all_idb_info = Features.export_idb_info('idb')
        all_pe_info = Features.export_pe_info('pe')
    else:
        print('convert_idb is error')

    # 만약 5개의 파일이 들어왔을때 그중 convert_idb 에러뜨는 것들은 제외시키고 나머지 것들만 다음 로직 수행되도록
    # 변경해야함.

    # 5. 분석 하기
    analyze = Analyze_files(all_idb_info, all_pe_info)

    yun_sorted_pe = dict()
    result_pe, yun_pe = analyze.analyze_pe()
    result_idb, yun_all = analyze.analyze_idb(yun_pe)

    yun_sorted_pe = sorted(yun_all.items(), key=lambda x: x[1]['timestamp_num'])

    #print(f"sorted_yun :: {json.dumps(yun_sorted_pe, indent=4)}")

    # 6. 결과 csv 저장 (임시)
    all_result = analyze.calculate_heuristic(result_idb, result_pe)

    #out_xlsx(r"C:\malware\result\test.xlsx", all_result)

    return all_result

if __name__ == "__main__":

    s = timeit.default_timer()

    start_engine()

    print(f"[+]time : {timeit.default_timer() - s}")


