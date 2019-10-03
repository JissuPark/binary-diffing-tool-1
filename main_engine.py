import hashlib
import json
import timeit
import os
import csv
import ssdeep
from multiprocessing import Process, Queue, Manager
from collections import OrderedDict
from Extract_Engine import pe2idb
from Extract_Engine.Flowchart_feature import extract_asm_and_const
from Extract_Engine.PE_feature import extract_pe
from Analzer_Engine import analyze_pe, analyze_flowchart
from openpyxl import load_workbook, Workbook



class Pe_Files_Check:
    '''
        * get_unique_pe_list
        -. idb로 변환할 PE 파일 중복 확인
        -. 파일명:해시값 딕셔너리에 저장 -> db or file
        -. 중복된 파일 제거
        -. Check the completely same PE file.
        -. make a dictionary of filename(key) and hashvalue(value)
        -. And then remove the same PE file.
        # exe_q에 idb로 변환할 exe파일을 쌓는다
        exe_q=Queue()
        * Return value
           dictionary of filename(key) and hashvalue(value)
        '''
    def __init__(self,pe_dir_path):
        self.pe_dir_path = pe_dir_path
        self.idat = 0
        self.idat64 = 1
        self.pe_check_error = -1
        self.pe_hash_dict = {}
        # STR_SIG_MZ = '0x4550'
        # HEX_M_32 = 0x14c
        # HEX_M_64 = 0x200

    def get_unique_pe_list(self):
        exe_list = os.listdir(self.pe_dir_path)
        for f in exe_list:
            f_path = os.path.join(self.pe_dir_path, f)
            f_hash = hashlib.sha256(open(f_path, 'rb').read()).hexdigest()

            # file hash 중복 = 완전히 같은 파일
            # 해당 파일은 삭제(이미 diffing할 동일 파일이 존재하므로)
            if f_hash in self.pe_hash_dict.values():
                os.remove(f_path)
            else:
                os.rename(f_path, os.path.join(self.pe_dir_path, f_hash))

                # 파일은 삭제하지만 해당 파일명(절대경로)와 해시정보는 DB에 있어야함.
            # 추후 시각화할 때 정보 필요
            self.pe_hash_dict[f_path] = f_hash


        # 이후에는 DB에 저장.
        # dictionary로 넘겨서 self.pe_hash_dict.value()의 유니크한 값들만 idb로 변환.
        # 나중에 서버에 올린 후에는 받은 파일 중 완전히 같은 파일은 서버 파일시스템에 저장하지 않는게 좋을 것 같다.
        # 예를 들어, file1, file2의 해시값이 동일한 경우
        # DB에는 file1과 file2의 파일명 그리고 해시값을 저장한다.
        # 그리고 서버 파일시스템 내에서 둘 중 하나의 파일은 지운다.
        # 일단 더 먼저 나오는 파일을 살리고 아닌 파일은 삭제하도록 하였다. 어차피 동일 파일이니깐.
        with open(r"C:\malware\result\test_pelist.txt", 'w') as pelist:
            json.dump(self.pe_hash_dict, pelist, ensure_ascii=False, indent='\t')

        return self.pe_hash_dict

def Convert_idb(PATH,IDB_PATH):
    # idb 변환
    return pe2idb.create_idb(PATH, IDB_PATH)

def multiprocess_file(q, return_dict, flag):

    while q.empty() != True:
        f_path = q.get()
        if flag == 'idb':
            info = extract_asm_and_const.basicblock_idb_info_extraction(f_path)  # 함수대표값 및 상수값 출력
        elif flag == 'pe':
            info = extract_pe.Pe_Feature(f_path).all()  # pe 속성 출력

        return_dict[f_path] = info


class Exract_Feature:
    def __init__(self, path, idb_path):
        self.path = path
        self.idb_path = idb_path

    def export_by_multi(self, flag):

        if flag == 'idb':
            path = self.idb_path
        elif flag == 'pe':
            path = self.path

        try:
            q = Queue()
            manager = Manager()
            pe2idb.exe_list_to_queue(path, q)
            return_dict = manager.dict()
            procs = list()
            for i in range(os.cpu_count() // 2 + 1):
                proc = Process(target=multiprocess_file, args=[q, return_dict, flag])
                procs.append(proc)
                proc.start()
            for p in procs:
                p.join()
            return return_dict
        except:
            return False

    def export_idb_info(self, flag):

        export_idb = self.export_by_multi(flag)

        if export_idb != False:
            count = 1
            for dict_list in export_idb.values():
                with open(r"C:\malware\result\idbfile_"+str(count)+".txt", 'w') as makefile:
                    json.dump(dict_list, makefile, ensure_ascii=False, indent='\t')
                count = count + 1
            return export_idb
        else:
            return False

    def export_pe_info(self, flag):

        export_pe = self.export_by_multi(flag)

        if export_pe != False:
            count = 1
            # print(return_dict)
            for dict_list in export_pe.values():
                with open(r"C:\malware\result\pefile_" + str(count) + ".txt", 'w') as makefile:
                    json.dump(dict_list, makefile, ensure_ascii=False, indent='\t')
                count = count + 1
            return export_pe
        else:
            return False

class Analyze_files:
    def __init__(self, all_idb_info, all_pe_info):
        self.all_pe_info = all_pe_info
        self.all_idb_info = all_idb_info

    def calculate_heuristic(self, idb_result, pe_result):
        '''
                가중치가 부여된 점수들을 더해서 반환해주는 함수
                *다 더했을 때 최대나 최소안에 있는지 확인하는 로직을 넣어주고 예외처리 해주면 될 듯
                :return: final score
                '''
        # 최종 휴리스틱 스코어

        real_final = OrderedDict()
        #semifinal = [0, 0, 0, 0, 0]
        pe_real_final = OrderedDict()

        #semifinal = [0,0,0,0,0]

        #final_score.append('0x1234')
        #Flowchart 점수 추가 (가중치 포함)

        for key_i, key_pe in zip(idb_result.items(), pe_result.items()):
            idb_final_score = OrderedDict()
            pe_final_score = OrderedDict()
            for value_i, value_pe in zip(key_i[1].items(), key_pe[1].items()):

                # semifinal = list()
                semifinal = [0, 0, 0, 0, 0]
                semifinal[1] = (value_i[1]['bbh'])
                semifinal[2] = (value_i[1]['const_value'])
                semifinal[3] = (value_pe[1]['imphash'])
                semifinal[4] = (value_pe[1]['rich'])

                idb_final_score[value_i[0]] = semifinal
                pe_final_score[value_pe[0]] = semifinal

            real_final[key_i[0]] = idb_final_score
            real_final[key_pe[0]] = pe_final_score


        return real_final

    def analyze_idb(self):
        idb = analyze_flowchart.AnalyzeFlowchart(self.all_idb_info)
        idb_split = idb.flow_parser()
        idb_result = idb.analyze_all(idb_split)
        return idb_result

    def analyze_pe(self):
        pe = analyze_pe.AnalyzePE(self.all_pe_info)
        pe_split = pe.pe_parser()

        pe_result = pe.analyze_all(pe_split)

        return pe_result

'''
    total score to the excel file
'''
def out_xlsx(path,result_dict):
    try:
        wb = load_workbook(path)
    except:
        wb = Workbook()
    ws = wb.create_sheet()
    ws = wb.active

    ws.title = 'result_xlsx'
    title = ['BASE_FILE', 'COMP_FILE', 'FILE HASH', 'BB HASH', 'CONSTANT', 'IMPORT HASH', 'RICH']
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

    for i in range(len(title)):
        ws[f'{cols[i]}1'] = title[i]
    target_count = len(result_dict) - 1
    start_row_num = 2

    for base, targets in result_dict.items():
        current_row_num = start_row_num
        ws[f'A{current_row_num}'] = base
        for t_name, t_infos in targets.items():
            ws[f'B{current_row_num}'] = t_name
            current_info = 0
            for t_info in t_infos:
                ws[f'{cols[current_info + 2]}{current_row_num}'] = t_info
                current_info += 1
            current_row_num += 1
        ws.merge_cells(f"A{start_row_num}:A{current_row_num - 1}")
        current_row_num += 1
        start_row_num = current_row_num

    #    wb.remove(wb['Sheet1'])
    wb.save(path)


if __name__ == "__main__":

    s = timeit.default_timer()

    PATH = r"C:\malware\mid_GandCrab_exe"
    IDB_PATH = r"C:\malware\mid_idb"

    pe_check = Pe_Files_Check(PATH)
    file_hash_dict = pe_check.get_unique_pe_list()
    flag = Convert_idb(PATH, IDB_PATH)
    Features = Exract_Feature(PATH, IDB_PATH)

    if flag == True:
        all_idb_info = Features.export_idb_info('idb')
        all_pe_info = Features.export_pe_info('pe')
    else:
        print('error fuck')
    print(type(all_idb_info))
    analyze = Analyze_files(all_idb_info, all_pe_info)

    result_idb = analyze.analyze_idb()
    # with open(r"C:\malware\result\idbtest.txt", 'w') as makefile:
    #     json.dump(result_idb, makefile, ensure_ascii=False, indent='\t')
    result_pe = analyze.analyze_pe()
    # with open(r"C:\malware\result\petest.txt", 'w') as makefile:
    #     json.dump(result_pe, makefile, ensure_ascii=False, indent='\t')

    all_result = analyze.calculate_heuristic(result_idb, result_pe)

    out_xlsx(r"C:\malware\result\test.xlsx", all_result)

#    out_csv(r"C:\malware\result\test.csv", all_result)

    print(f"[+]time : {timeit.default_timer() - s}")

    ########################### 최종 결과물 csv 추출 ###################################

    # dict = {
    #           "file1" : {
    #                        "file2":["0x456",1, 2, 3, 4, 5, 6],
    #                        "file3":["0x789",1, 2, 3, 4, 5, 6],
    #                        "file4":["0x012",1, 2, 3, 4, 5, 6],
    #                        "file5":["0x345",1, 2, 3, 4, 5, 6]
    #                     }
    #           "file2" : {
    #                        "file1":["0x456",1, 2, 3, 4, 5, 6],
    #                        "file3":["0x789",1, 2, 3, 4, 5, 6],
    #                        "file4":["0x012",1, 2, 3, 4, 5, 6],
    #                        "file5":["0x345",1, 2, 3, 4, 5, 6]
    #                     }
    #           "file3" : {
    #                        "file1":["0x456",1, 2, 3, 4, 5, 6],
    #                        "file2":["0x789",1, 2, 3, 4, 5, 6],
    #                        "file4":["0x012",1, 2, 3, 4, 5, 6],
    #                        "file5":["0x345",1, 2, 3, 4, 5, 6]
    #                     }
    #           "file4" : {
    #                        "file1":["0x456",1, 2, 3, 4, 5, 6],
    #                        "file2":["0x789",1, 2, 3, 4, 5, 6],
    #                        "file3":["0x012",1, 2, 3, 4, 5, 6],
    #                        "file5":["0x345",1, 2, 3, 4, 5, 6]
    #                     }
    #           "file5" : {
    #                        "file2":["0x456",1, 2, 3, 4, 5, 6],
    #                        "file3":["0x789",1, 2, 3, 4, 5, 6],
    #                        "file4":["0x012",1, 2, 3, 4, 5, 6],
    #                        "file5":["0x345",1, 2, 3, 4, 5, 6]
    #                     }
    #        }